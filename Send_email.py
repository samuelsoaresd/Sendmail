import smtplib, ssl
import openpyxl
from email.message import EmailMessage

file_path = ""  #Caminho do arquivo xls
wb = openpyxl.load_workbook(file_path)
sheet = wb["Planilha1"]

emails_column = 1  # Coluna que contém os endereços de e-mail
status_column = 2  # Coluna onde o status de envio será registrado

#Layout html do e-mail 
html = """<!doctype html>   
</html>
"""

# Cria um contexto SSL seguro
context = ssl.create_default_context()

with smtplib.SMTP_SSL('smtp.gmail.com', 465, context=context) as smtp:
  smtp.login('email gmail', 'Senha app gmail') #Endereço de email para envios
  try:
    for row in range(2, sheet.max_row + 1): #começa pela segunda linha por conta do cabeçalho
      email = sheet.cell(row=row, column=emails_column).value
      status = sheet.cell(row=row, column=status_column).value
      if email and not status:
        msg = EmailMessage()
        msg.set_content('Seja bem vindo') 
        msg.add_alternative(html, subtype='html')
        msg['Subject'] = ''
        msg['From'] = 'email'
        msg['To'] = email
        msg['Cc'] = ''
        msg['Bcc'] = ''
        smtp.send_message(msg)
        sheet.cell(row=row, column=status_column).value = 'Enviado'
  except Exception as e:
      print("Erro ao enviar e-mail:", e)

wb.save(file_path)  # Salva as alterações no arquivo xls

print("E-mails enviados")
